#!/usr/bin/env python3
"""
Excel to PDF Converter (Native)

Uses native system tools to convert Excel to PDF, preserving original formatting.
This is much simpler and more accurate than custom rendering.

Supports:
- macOS: Uses LibreOffice or Excel (via AppleScript)
- Linux: Uses LibreOffice (unoconv or headless)
- Windows: Uses Excel automation (win32com)

Usage:
    python excel_to_pdf_native.py input.xlsx output.pdf
    python excel_to_pdf_native.py input.xlsx output.pdf --sheet "Sheet1"
    python excel_to_pdf_native.py input.xlsx output_dir/ --all-sheets
"""

import argparse
import sys
import platform
import subprocess
import shutil
from pathlib import Path
from typing import Optional
import openpyxl
from logging_utils import setup_script_logging, log_operation, log_error_with_context

# Setup logging
logger = setup_script_logging("excel_to_pdf_native")


def check_libreoffice() -> Optional[str]:
    """Check if LibreOffice is available and return path."""
    # Common LibreOffice paths
    possible_paths = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",  # macOS
        "/usr/bin/soffice",  # Linux
        "/usr/bin/libreoffice",  # Linux alternative
        "C:\\Program Files\\LibreOffice\\program\\soffice.exe",  # Windows
    ]

    for path in possible_paths:
        if Path(path).exists():
            logger.info("Found LibreOffice at: %s", path)
            return path

    # Try to find in PATH
    libreoffice = shutil.which("soffice") or shutil.which("libreoffice")
    if libreoffice:
        logger.info("Found LibreOffice in PATH: %s", libreoffice)
        return libreoffice

    return None


def convert_with_libreoffice(
    input_path: Path, output_path: Path, sheet_name: Optional[str] = None
) -> bool:
    """
    Convert Excel to PDF using LibreOffice headless mode.

    Args:
        input_path: Excel file path
        output_path: Output PDF path
        sheet_name: Optional sheet name (LibreOffice converts all sheets)

    Returns:
        True if successful
    """
    soffice = check_libreoffice()
    if not soffice:
        logger.error("LibreOffice not found")
        return False

    try:
        # LibreOffice headless conversion
        output_dir = output_path.parent
        output_dir.mkdir(parents=True, exist_ok=True)

        cmd = [
            soffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(output_dir),
            str(input_path),
        ]

        logger.info("Running: %s", " ".join(cmd))
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60, check=False)

        if not result.returncode:
            # LibreOffice creates output with same name as input
            generated_pdf = output_dir / f"{input_path.stem}.pdf"

            if generated_pdf.exists():
                # Rename if needed
                if generated_pdf != output_path:
                    generated_pdf.rename(output_path)

                logger.info("Successfully converted to: %s", output_path)
                print(f"✓ Created: {output_path}")
                return True
            else:
                logger.error("PDF not created: %s", generated_pdf)
                return False

        logger.error("LibreOffice error: %s", result.stderr)
        print(f"\u2717 Error: {result.stderr}", file=sys.stderr)
        return False

    except subprocess.TimeoutExpired:
        logger.error("LibreOffice conversion timed out")
        print("✗ Error: Conversion timed out", file=sys.stderr)
        return False
    except Exception as e:
        log_error_with_context(logger, e, {"input": str(input_path)})
        print(f"✗ Error: {e}", file=sys.stderr)
        return False


def convert_with_excel_macos(
    input_path: Path, output_path: Path, sheet_name: Optional[str] = None
) -> bool:
    """
    Convert Excel to PDF using Excel on macOS (via AppleScript).

    Args:
        input_path: Excel file path
        output_path: Output PDF path
        sheet_name: Optional sheet name to export

    Returns:
        True if successful
    """
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # AppleScript to open Excel and save as PDF
        applescript = f"""
        tell application "Microsoft Excel"
            open POSIX file "{input_path.absolute()}"
            set workbook to active workbook
            
            {f'set active sheet to worksheet "{sheet_name}" of workbook' if sheet_name else ''}
            
            save workbook as PDF in POSIX file "{output_path.absolute()}"
            close workbook saving no
        end tell
        """

        logger.info("Using Excel for macOS to convert")
        result = subprocess.run(
            ["osascript", "-e", applescript],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )

        if not result.returncode and output_path.exists():
            logger.info("Successfully converted with Excel: %s", output_path)
            print(f"✓ Created: {output_path}")
            return True

        logger.error("Excel conversion failed: %s", result.stderr)
        return False

    except Exception as e:
        log_error_with_context(logger, e, {"input": str(input_path)})
        return False


def extract_sheet_to_temp_file(workbook_path: Path, sheet_name: str) -> Optional[Path]:
    """
    Extract a single sheet to a temporary Excel file.

    Args:
        workbook_path: Original workbook path
        sheet_name: Sheet to extract

    Returns:
        Path to temporary file with single sheet
    """
    try:
        wb = openpyxl.load_workbook(workbook_path)

        if sheet_name not in wb.sheetnames:
            logger.error("Sheet '%s' not found", sheet_name)
            return None

        # Create new workbook with only the target sheet
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet

        # Copy the sheet
        source_sheet = wb[sheet_name]
        target_sheet = new_wb.create_sheet(sheet_name)

        # Copy cell values and styles
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet[cell.coordinate]
                target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = cell.alignment.copy()

        # Copy column widths
        for col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[
                col_letter
            ].width

        # Save to temp file
        temp_file = workbook_path.parent / f"_temp_{sheet_name.replace(' ', '_')}.xlsx"
        new_wb.save(temp_file)

        logger.debug("Created temp file: %s", temp_file)
        return temp_file

    except Exception as e:
        log_error_with_context(logger, e, {"sheet": sheet_name})
        return None


def convert_excel_to_pdf(
    input_path: Path, output_path: Path, sheet_name: Optional[str] = None, method: str = "auto"
) -> bool:
    """
    Convert Excel to PDF using best available method.

    Args:
        input_path: Excel file path
        output_path: Output PDF path
        sheet_name: Optional sheet name (if None, converts all)
        method: 'auto', 'libreoffice', 'excel', or 'fallback'

    Returns:
        True if successful
    """
    temp_file = None

    try:
        # If single sheet requested, create temp file with just that sheet
        if sheet_name:
            wb = openpyxl.load_workbook(input_path)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

            temp_file = extract_sheet_to_temp_file(input_path, sheet_name)
            if not temp_file:
                return False
            conversion_path = temp_file
        else:
            conversion_path = input_path

        # Try methods in order
        if method == "auto":
            system = platform.system()

            # Try LibreOffice first (most reliable cross-platform)
            if convert_with_libreoffice(conversion_path, output_path, sheet_name):
                log_operation(
                    logger,
                    "converted",
                    input=str(input_path),
                    output=str(output_path),
                    sheet=sheet_name,
                    method="libreoffice",
                )
                return True

            # Try Excel on macOS
            if system == "Darwin":
                if convert_with_excel_macos(conversion_path, output_path, sheet_name):
                    log_operation(
                        logger,
                        "converted",
                        input=str(input_path),
                        output=str(output_path),
                        sheet=sheet_name,
                        method="excel_macos",
                    )
                    return True

            # Fallback to custom renderer (previous implementation)
            logger.warning("Native conversion methods failed, falling back to custom renderer")
            print("⚠️  Native tools not available, using fallback renderer")
            return False

        elif method == "libreoffice":
            return convert_with_libreoffice(conversion_path, output_path, sheet_name)

        elif method == "excel":
            if platform.system() == "Darwin":
                return convert_with_excel_macos(conversion_path, output_path, sheet_name)
            else:
                logger.error("Excel automation only supported on macOS")
                return False

        return False

    finally:
        # Clean up temp file
        if temp_file and temp_file.exists():
            temp_file.unlink()
            logger.debug("Cleaned up temp file: %s", temp_file)


def main():
    """Command-line interface."""
    parser = argparse.ArgumentParser(
        description="Convert Excel to PDF using native tools",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert entire workbook
  python excel_to_pdf_native.py data.xlsx output.pdf
  
  # Convert single sheet
  python excel_to_pdf_native.py data.xlsx output.pdf --sheet "Sheet1"
  
  # Convert all sheets to separate PDFs
  python excel_to_pdf_native.py data.xlsx output_dir/ --all-sheets
  
  # Specify method
  python excel_to_pdf_native.py data.xlsx output.pdf --method libreoffice
  
Requirements:
  - LibreOffice (recommended): brew install --cask libreoffice
  - Or Microsoft Excel (macOS only)
        """,
    )

    parser.add_argument("input", type=Path, help="Input Excel file")
    parser.add_argument("output", type=Path, nargs="?", help="Output PDF file or directory")
    parser.add_argument("--sheet", type=str, help="Sheet name to convert")
    parser.add_argument("--all-sheets", action="store_true", help="Convert all sheets separately")
    parser.add_argument(
        "--method",
        choices=["auto", "libreoffice", "excel"],
        default="auto",
        help="Conversion method",
    )
    parser.add_argument("--list", action="store_true", help="List sheets and exit")

    args = parser.parse_args()

    # Validate input
    if not args.input.exists():
        print(f"Error: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    # List sheets
    if args.list:
        wb = openpyxl.load_workbook(args.input)
        print(f"\n📊 Sheets in {args.input.name}:")
        print("=" * 50)
        for i, sheet in enumerate(wb.sheetnames, 1):
            print(f"{i}. {sheet}")
        print("=" * 50)
        print(f"Total: {len(wb.sheetnames)} sheets\n")
        return

    # Check for LibreOffice
    if not check_libreoffice() and platform.system() != "Darwin":
        print("\n⚠️  Warning: LibreOffice not found!")
        print("Install with: brew install --cask libreoffice")
        print("Or: sudo apt-get install libreoffice\n")

    try:
        # Convert all sheets
        if args.all_sheets:
            wb = openpyxl.load_workbook(args.input)
            output_dir = (
                args.output if not args.output.suffix else args.output.parent / args.output.stem
            )
            output_dir.mkdir(parents=True, exist_ok=True)

            print(f"\n📊 Converting all sheets from {args.input.name}")
            print("=" * 60)

            success_count = 0
            for sheet_name in wb.sheetnames:
                safe_name = "".join(
                    c if c.isalnum() or c in (" ", "-", "_") else "_" for c in sheet_name
                )
                output_path = output_dir / f"{safe_name}.pdf"

                if convert_excel_to_pdf(args.input, output_path, sheet_name, args.method):
                    success_count += 1

            print("=" * 60)
            print(f"✓ Successfully converted {success_count}/{len(wb.sheetnames)} sheets\n")

        # Convert single sheet or entire workbook
        else:
            sheet_desc = f"sheet '{args.sheet}'" if args.sheet else "workbook"
            print(f"\n📊 Converting {sheet_desc} from {args.input.name}")
            print("=" * 60)

            if convert_excel_to_pdf(args.input, args.output, args.sheet, args.method):
                print("=" * 60)
                print(f"✓ Successfully converted to {args.output}\n")
            else:
                print("=" * 60)
                print("✗ Conversion failed\n", file=sys.stderr)
                sys.exit(1)

    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        log_error_with_context(logger, e, {"input": str(args.input)})
        sys.exit(1)


if __name__ == "__main__":
    main()
