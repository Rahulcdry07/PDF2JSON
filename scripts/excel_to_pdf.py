#!/usr/bin/env python3
"""
Excel to PDF Converter

Extracts individual sheets from Excel files (.xlsx, .xls) and converts them to PDF.
Supports multiple sheets, custom formatting, and batch processing.

Usage:
    python excel_to_pdf.py input.xlsx output.pdf --sheet "Sheet1"
    python excel_to_pdf.py input.xlsx output_dir/ --all-sheets
    python excel_to_pdf.py input.xlsx output.pdf --sheets "Sheet1,Sheet2"
"""

import argparse
import sys
from pathlib import Path
from typing import List
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from logging_utils import setup_script_logging, log_operation

# Setup logging
logger = setup_script_logging("excel_to_pdf")


class ExcelToPDFConverter:
    """Convert Excel sheets to PDF with formatting."""

    def __init__(self, excel_path: Path):
        """
        Initialize converter with Excel file.

        Args:
            excel_path: Path to Excel file
        """
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            logger.error(f"Excel file not found: {excel_path}")
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        logger.info(f"Initializing converter for: {self.excel_path.name}")
        self.workbook = openpyxl.load_workbook(excel_path, data_only=True)
        logger.debug(f"Loaded workbook with {len(self.workbook.sheetnames)} sheets")

    def list_sheets(self) -> List[str]:
        """Get list of all sheet names in the workbook."""
        return self.workbook.sheetnames

    def get_sheet_data(self, sheet_name: str) -> List[List]:
        """
        Extract data from a specific sheet.

        Args:
            sheet_name: Name of the sheet to extract

        Returns:
            2D list of cell values
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.list_sheets()}")

        sheet = self.workbook[sheet_name]
        data = []

        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ""
                row_data.append(str(value))
            data.append(row_data)

        # Remove empty trailing rows
        while data and all(cell == "" for cell in data[-1]):
            data.pop()

        return data

    def convert_sheet_to_pdf(
        self,
        sheet_name: str,
        output_path: Path,
        page_size: tuple = A4,
        orientation: str = "portrait",
        fit_to_width: bool = True,
    ) -> None:
        """
        Convert a single sheet to PDF.

        Args:
            sheet_name: Name of sheet to convert
            output_path: Output PDF path
            page_size: Page size tuple (default: A4)
            orientation: 'portrait' or 'landscape'
            fit_to_width: Automatically adjust column widths
        """
        data = self.get_sheet_data(sheet_name)

        if not data:
            raise ValueError(f"Sheet '{sheet_name}' is empty")

        # Auto-detect orientation for wide tables
        num_cols = len(data[0]) if data else 1
        if orientation == "auto":
            # Use landscape for tables with 6+ columns
            orientation = "landscape" if num_cols >= 6 else "portrait"

        # Adjust page size for orientation
        if orientation == "landscape":
            page_size = landscape(page_size)

        # Standard margins - wider tables will span multiple pages
        left_margin = 0.5 * inch
        right_margin = 0.5 * inch
        top_margin = 0.7 * inch
        bottom_margin = 0.7 * inch

        if num_cols >= 12:
            logger.warning(f"Table has {num_cols} columns - may span multiple pages")
            print(f"⚠️  Table has {num_cols} columns. Using multi-page layout for readability.")

        # Create PDF
        output_path.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=page_size,
            leftMargin=left_margin,
            rightMargin=right_margin,
            topMargin=top_margin,
            bottomMargin=bottom_margin,
        )

        # Build story
        story = []
        styles = getSampleStyleSheet()

        # Add title
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Heading1"], fontSize=14, textColor=colors.HexColor("#667eea"), spaceAfter=8
        )
        story.append(Paragraph(f"<b>{sheet_name}</b>", title_style))
        story.append(Spacer(1, 0.15 * inch))

        # Calculate column widths based on content with readable minimums
        page_width = page_size[0] - left_margin - right_margin

        col_widths = []
        for col_idx in range(num_cols):
            max_len = 0
            for row in data[:30]:  # Sample more rows for accuracy
                if col_idx < len(row):
                    cell_content = str(row[col_idx])
                    max_len = max(max_len, len(cell_content))

            # Set minimum readable widths - don't go below these
            if num_cols >= 15:
                # Very wide: much larger minimum for readability
                min_width = 1.8 * inch
                max_width = 4.5 * inch
                char_width = 8
            elif num_cols >= 10:
                min_width = 2.0 * inch
                max_width = 5.0 * inch
                char_width = 9
            else:
                min_width = 2.2 * inch
                max_width = 5.5 * inch
                char_width = 10

            # Calculate width with readable constraints
            width = min(max(min_width, max_len * char_width), max_width)
            col_widths.append(width)

        total_width = sum(col_widths)

        # Only scale down if slightly over, otherwise let it span pages
        if total_width > page_width and total_width <= page_width * 1.15:
            # Minor adjustment - scale to fit
            scale = page_width / total_width
            col_widths = [w * scale for w in col_widths]
            logger.debug(f"Scaled columns by {scale:.2f} to fit page width")
        elif total_width > page_width:
            # Table too wide - will span multiple pages horizontally
            logger.info(f"Table width ({total_width:.0f}pt) exceeds page ({page_width:.0f}pt) - will span pages")
            print("📄 Table spans multiple pages for better readability")

        # Use larger, more readable font sizes
        if num_cols >= 15:
            header_font_size = 11
            data_font_size = 10
            padding = 8
        elif num_cols >= 10:
            header_font_size = 12
            data_font_size = 11
            padding = 9
        else:
            header_font_size = 13
            data_font_size = 12
            padding = 10

        logger.info(f"Using font sizes: header={header_font_size}pt, data={data_font_size}pt")

        # Create table with repeatRows for header on each page
        table = Table(data, colWidths=col_widths, repeatRows=1, splitByRow=True)

        # Apply clean, readable table style
        table_style = TableStyle(
            [
                # Header row
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), header_font_size),
                ("BOTTOMPADDING", (0, 0), (-1, 0), padding + 4),
                ("TOPPADDING", (0, 0), (-1, 0), padding + 4),
                # Data rows
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), data_font_size),
                ("GRID", (0, 0), (-1, -1), 0.75, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), padding + 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), padding + 2),
                ("TOPPADDING", (0, 1), (-1, -1), padding + 1),
                ("BOTTOMPADDING", (0, 1), (-1, -1), padding + 1),
                # Alternating row colors for readability
                *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f7f7f7")) for i in range(2, len(data), 2)],
            ]
        )

        table.setStyle(table_style)
        story.append(table)

        # Build PDF
        doc.build(story)
        log_operation(
            logger,
            "sheet_converted",
            sheet=sheet_name,
            output=str(output_path),
            rows=len(data),
            cols=num_cols,
            orientation=orientation,
        )
        print(f"✓ Converted sheet '{sheet_name}' to {output_path}")

    def convert_all_sheets(self, output_dir: Path, page_size: tuple = A4, orientation: str = "portrait") -> List[Path]:
        """
        Convert all sheets to separate PDF files.

        Args:
            output_dir: Output directory for PDFs
            page_size: Page size tuple
            orientation: 'portrait' or 'landscape'

        Returns:
            List of created PDF paths
        """
        output_dir.mkdir(parents=True, exist_ok=True)
        created_files = []

        logger.info(f"Converting all sheets ({len(self.list_sheets())}) to separate PDFs")

        for sheet_name in self.list_sheets():
            safe_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in sheet_name)
            output_path = output_dir / f"{safe_name}.pdf"

            try:
                self.convert_sheet_to_pdf(sheet_name, output_path, page_size, orientation)
                created_files.append(output_path)
            except Exception as e:
                print(f"✗ Error converting sheet '{sheet_name}': {e}", file=sys.stderr)

        return created_files

    def convert_multiple_sheets(
        self, sheet_names: List[str], output_path: Path, page_size: tuple = A4, orientation: str = "portrait"
    ) -> None:
        """
        Convert multiple sheets to a single PDF.

        Args:
            sheet_names: List of sheet names to convert
            output_path: Output PDF path
            page_size: Page size tuple
            orientation: 'portrait' or 'landscape'
        """
        if orientation == "landscape":
            page_size = landscape(page_size)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=page_size,
            leftMargin=0.5 * inch,
            rightMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        story = []
        styles = getSampleStyleSheet()

        for idx, sheet_name in enumerate(sheet_names):
            if idx > 0:
                story.append(PageBreak())

            try:
                data = self.get_sheet_data(sheet_name)

                # Add sheet title
                title_style = ParagraphStyle(
                    "CustomTitle",
                    parent=styles["Heading1"],
                    fontSize=16,
                    textColor=colors.HexColor("#667eea"),
                    spaceAfter=12,
                )
                story.append(Paragraph(f"<b>{sheet_name}</b>", title_style))
                story.append(Spacer(1, 0.2 * inch))

                # Create table
                page_width = page_size[0] - doc.leftMargin - doc.rightMargin
                num_cols = len(data[0]) if data else 1
                col_width = page_width / num_cols
                col_widths = [col_width] * num_cols

                table = Table(data, colWidths=col_widths, repeatRows=1)

                # Apply style
                table_style = TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("TOPPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 1), (-1, -1), 6),
                        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                        *[("BACKGROUND", (0, i), (-1, i), colors.HexColor("#f9f9f9")) for i in range(2, len(data), 2)],
                    ]
                )

                table.setStyle(table_style)
                story.append(table)

                print(f"✓ Added sheet '{sheet_name}'")

            except Exception as e:
                print(f"✗ Error processing sheet '{sheet_name}': {e}", file=sys.stderr)

        doc.build(story)
        print(f"✓ Created combined PDF: {output_path}")

    def close(self):
        """Close the workbook."""
        self.workbook.close()


def main():
    """Command-line interface."""
    parser = argparse.ArgumentParser(
        description="Convert Excel sheets to PDF",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert single sheet
  python excel_to_pdf.py data.xlsx output.pdf --sheet "Sheet1"
  
  # Convert all sheets to separate PDFs
  python excel_to_pdf.py data.xlsx output_dir/ --all-sheets
  
  # Convert multiple sheets to one PDF
  python excel_to_pdf.py data.xlsx combined.pdf --sheets "Sheet1,Sheet2,Sheet3"
  
  # List all sheets
  python excel_to_pdf.py data.xlsx --list
  
  # Use landscape orientation
  python excel_to_pdf.py data.xlsx output.pdf --sheet "Sheet1" --landscape
        """,
    )

    parser.add_argument("input", type=Path, help="Input Excel file (.xlsx, .xls)")
    parser.add_argument("output", type=Path, nargs="?", help="Output PDF file or directory")
    parser.add_argument("--sheet", type=str, help="Sheet name to convert")
    parser.add_argument("--sheets", type=str, help="Comma-separated sheet names")
    parser.add_argument("--all-sheets", action="store_true", help="Convert all sheets")
    parser.add_argument("--list", action="store_true", help="List all sheets and exit")
    parser.add_argument("--landscape", action="store_true", help="Use landscape orientation")
    parser.add_argument("--page-size", choices=["A4", "Letter"], default="A4", help="Page size")

    args = parser.parse_args()

    # Validate input
    if not args.input.exists():
        print(f"Error: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    if not args.input.suffix.lower() in [".xlsx", ".xls"]:
        print("Error: Invalid file type. Expected .xlsx or .xls", file=sys.stderr)
        sys.exit(1)

    # Create converter
    converter = ExcelToPDFConverter(args.input)

    # List sheets mode
    if args.list:
        sheets = converter.list_sheets()
        print(f"\n📊 Sheets in {args.input.name}:")
        print("=" * 50)
        for i, sheet in enumerate(sheets, 1):
            print(f"{i}. {sheet}")
        print("=" * 50)
        print(f"Total: {len(sheets)} sheets\n")
        converter.close()
        return

    # Validate output
    if not args.output:
        print("Error: Output path required", file=sys.stderr)
        parser.print_help()
        sys.exit(1)

    # Determine page size
    page_size = letter if args.page_size == "Letter" else A4
    orientation = "landscape" if args.landscape else "portrait"

    try:
        # Convert all sheets
        if args.all_sheets:
            if not args.output.suffix:
                output_dir = args.output
            else:
                output_dir = args.output.parent / args.output.stem

            print(f"\n📊 Converting all sheets from {args.input.name}")
            print("=" * 60)
            created = converter.convert_all_sheets(output_dir, page_size, orientation)
            print("=" * 60)
            print(f"✓ Created {len(created)} PDF files in {output_dir}\n")

        # Convert multiple sheets to one PDF
        elif args.sheets:
            sheet_list = [s.strip() for s in args.sheets.split(",")]
            print(f"\n📊 Converting {len(sheet_list)} sheets to {args.output.name}")
            print("=" * 60)
            converter.convert_multiple_sheets(sheet_list, args.output, page_size, orientation)
            print("=" * 60)
            print("✓ Done!\n")

        # Convert single sheet
        elif args.sheet:
            print(f"\n📊 Converting sheet '{args.sheet}' to {args.output.name}")
            print("=" * 60)
            converter.convert_sheet_to_pdf(args.sheet, args.output, page_size, orientation)
            print("=" * 60)
            print("✓ Done!\n")

        else:
            print("Error: Specify --sheet, --sheets, or --all-sheets", file=sys.stderr)
            parser.print_help()
            sys.exit(1)

    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        converter.close()


if __name__ == "__main__":
    main()
