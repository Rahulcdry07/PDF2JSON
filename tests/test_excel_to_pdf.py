#!/usr/bin/env python3
"""Tests for excel_to_pdf.py script."""

import sys
from pathlib import Path
import pytest
from openpyxl import Workbook

# Add scripts directory to path
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from excel_to_pdf import ExcelToPDFConverter


@pytest.fixture
def sample_excel_file(tmp_path):
    """Create a sample Excel file for testing."""
    excel_file = tmp_path / "test.xlsx"
    wb = Workbook()

    # First sheet with data
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Name"
    ws1["B1"] = "Age"
    ws1["C1"] = "City"
    ws1["A2"] = "Alice"
    ws1["B2"] = 30
    ws1["C2"] = "NYC"
    ws1["A3"] = "Bob"
    ws1["B3"] = 25
    ws1["C3"] = "LA"

    # Second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Product"
    ws2["B1"] = "Price"
    ws2["A2"] = "Apple"
    ws2["B2"] = 1.5

    # Empty sheet
    wb.create_sheet("EmptySheet")

    wb.save(excel_file)
    return excel_file


def test_converter_init(sample_excel_file):
    """Test ExcelToPDFConverter initialization."""
    converter = ExcelToPDFConverter(sample_excel_file)
    assert converter.excel_path == sample_excel_file
    assert converter.workbook is not None


def test_converter_init_file_not_found():
    """Test initialization with non-existent file."""
    with pytest.raises(FileNotFoundError):
        ExcelToPDFConverter(Path("nonexistent.xlsx"))


def test_list_sheets(sample_excel_file):
    """Test listing all sheets."""
    converter = ExcelToPDFConverter(sample_excel_file)
    sheets = converter.list_sheets()

    assert "Sheet1" in sheets
    assert "Sheet2" in sheets
    assert "EmptySheet" in sheets
    assert len(sheets) == 3


def test_get_sheet_data(sample_excel_file):
    """Test extracting sheet data."""
    converter = ExcelToPDFConverter(sample_excel_file)
    data = converter.get_sheet_data("Sheet1")

    # Should have 3 rows
    assert len(data) == 3

    # Check header row
    assert data[0] == ["Name", "Age", "City"]

    # Check data rows
    assert data[1] == ["Alice", "30", "NYC"]
    assert data[2] == ["Bob", "25", "LA"]


def test_get_sheet_data_invalid_sheet(sample_excel_file):
    """Test getting data from non-existent sheet."""
    converter = ExcelToPDFConverter(sample_excel_file)

    with pytest.raises(ValueError, match="Sheet 'InvalidSheet' not found"):
        converter.get_sheet_data("InvalidSheet")


def test_get_sheet_data_empty_cells(sample_excel_file):
    """Test that empty cells are converted to empty strings."""
    converter = ExcelToPDFConverter(sample_excel_file)
    data = converter.get_sheet_data("Sheet2")

    # All cells should be strings
    for row in data:
        for cell in row:
            assert isinstance(cell, str)


def test_convert_sheet_to_pdf(sample_excel_file, tmp_path):
    """Test converting a sheet to PDF."""
    converter = ExcelToPDFConverter(sample_excel_file)
    output_pdf = tmp_path / "output.pdf"

    converter.convert_sheet_to_pdf("Sheet1", output_pdf)

    # PDF should be created
    assert output_pdf.exists()
    assert output_pdf.stat().st_size > 0


def test_convert_sheet_to_pdf_landscape(sample_excel_file, tmp_path):
    """Test converting with landscape orientation."""
    converter = ExcelToPDFConverter(sample_excel_file)
    output_pdf = tmp_path / "landscape.pdf"

    converter.convert_sheet_to_pdf("Sheet1", output_pdf, orientation="landscape")

    assert output_pdf.exists()


def test_convert_sheet_to_pdf_auto_orientation(sample_excel_file, tmp_path):
    """Test auto orientation detection."""
    converter = ExcelToPDFConverter(sample_excel_file)
    output_pdf = tmp_path / "auto.pdf"

    # Sheet1 has 3 columns, should be portrait
    converter.convert_sheet_to_pdf("Sheet1", output_pdf, orientation="auto")

    assert output_pdf.exists()


def test_convert_empty_sheet_raises_error(sample_excel_file, tmp_path):
    """Test that converting empty sheet raises error."""
    converter = ExcelToPDFConverter(sample_excel_file)
    output_pdf = tmp_path / "empty.pdf"

    with pytest.raises(ValueError, match="Sheet 'EmptySheet' is empty"):
        converter.convert_sheet_to_pdf("EmptySheet", output_pdf)


def test_convert_creates_parent_directories(sample_excel_file, tmp_path):
    """Test that PDF creation creates parent directories."""
    converter = ExcelToPDFConverter(sample_excel_file)
    output_pdf = tmp_path / "nested" / "dir" / "output.pdf"

    converter.convert_sheet_to_pdf("Sheet1", output_pdf)

    assert output_pdf.exists()
    assert output_pdf.parent.exists()


def test_get_sheet_data_trailing_empty_rows(tmp_path):
    """Test that trailing empty rows are removed."""
    excel_file = tmp_path / "trailing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Data"
    ws["A2"] = "More"
    # Row 3-5 will be empty
    wb.save(excel_file)

    converter = ExcelToPDFConverter(excel_file)
    data = converter.get_sheet_data(ws.title)

    # Should only have 2 rows (empty trailing rows removed)
    assert len(data) == 2


def test_convert_sheet_with_numbers(tmp_path):
    """Test converting sheet with various data types."""
    excel_file = tmp_path / "numbers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Integer"
    ws["B1"] = "Float"
    ws["C1"] = "Text"
    ws["A2"] = 42
    ws["B2"] = 3.14
    ws["C2"] = "Hello"
    wb.save(excel_file)

    converter = ExcelToPDFConverter(excel_file)
    data = converter.get_sheet_data(ws.title)

    # All should be converted to strings
    assert data[1][0] == "42"
    assert data[1][1] == "3.14"
    assert data[1][2] == "Hello"

    # Can also convert to PDF
    output_pdf = tmp_path / "numbers.pdf"
    converter.convert_sheet_to_pdf(ws.title, output_pdf)
    assert output_pdf.exists()
