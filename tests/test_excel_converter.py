"""
Tests for Excel to PDF converter functionality.
"""

import pytest
from pathlib import Path
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill

# Add scripts to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

from excel_to_pdf import ExcelToPDFConverter


@pytest.fixture
def sample_excel(tmp_path):
    """Create a sample Excel file for testing."""
    excel_file = tmp_path / "test.xlsx"

    wb = openpyxl.Workbook()

    # Sheet 1: Simple data
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Code", "Description", "Unit", "Rate"])
    ws1.append(["1.1", "Excavation work", "cum", "250.50"])
    ws1.append(["1.2", "Concrete work", "cum", "5000.00"])
    ws1.append(["1.3", "Steel work", "kg", "75.25"])

    # Sheet 2: More data
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Item", "Quantity", "Price"])
    ws2.append(["Item A", "10", "100"])
    ws2.append(["Item B", "20", "200"])

    # Sheet 3: Empty sheet
    wb.create_sheet("EmptySheet")

    wb.save(excel_file)
    return excel_file


def test_converter_initialization(sample_excel):
    """Test converter initialization."""
    converter = ExcelToPDFConverter(sample_excel)
    assert converter.excel_path == sample_excel
    assert converter.workbook is not None
    converter.close()


def test_converter_nonexistent_file():
    """Test converter with nonexistent file."""
    with pytest.raises(FileNotFoundError):
        ExcelToPDFConverter(Path("nonexistent.xlsx"))


def test_list_sheets(sample_excel):
    """Test listing sheets."""
    converter = ExcelToPDFConverter(sample_excel)
    sheets = converter.list_sheets()

    assert "Sheet1" in sheets
    assert "Sheet2" in sheets
    assert "EmptySheet" in sheets
    assert len(sheets) == 3

    converter.close()


def test_get_sheet_data(sample_excel):
    """Test extracting sheet data."""
    converter = ExcelToPDFConverter(sample_excel)

    data = converter.get_sheet_data("Sheet1")
    assert len(data) == 4  # Header + 3 rows
    assert data[0] == ["Code", "Description", "Unit", "Rate"]
    assert data[1][0] == "1.1"
    assert data[1][3] == "250.50"  # Excel preserves decimal places

    converter.close()


def test_get_sheet_data_invalid_sheet(sample_excel):
    """Test getting data from invalid sheet."""
    converter = ExcelToPDFConverter(sample_excel)

    with pytest.raises(ValueError, match="Sheet 'Invalid' not found"):
        converter.get_sheet_data("Invalid")

    converter.close()


def test_get_empty_sheet_data(sample_excel):
    """Test getting data from empty sheet."""
    converter = ExcelToPDFConverter(sample_excel)

    data = converter.get_sheet_data("EmptySheet")
    assert len(data) == 0

    converter.close()


def test_convert_single_sheet(sample_excel, tmp_path):
    """Test converting a single sheet to PDF."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "output.pdf"

    converter.convert_sheet_to_pdf("Sheet1", output_pdf)

    assert output_pdf.exists()
    assert output_pdf.stat().st_size > 0

    converter.close()


def test_convert_sheet_portrait(sample_excel, tmp_path):
    """Test converting sheet with portrait orientation."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "portrait.pdf"

    from reportlab.lib.pagesizes import A4

    converter.convert_sheet_to_pdf("Sheet1", output_pdf, page_size=A4, orientation="portrait")

    assert output_pdf.exists()
    converter.close()


def test_convert_sheet_landscape(sample_excel, tmp_path):
    """Test converting sheet with landscape orientation."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "landscape.pdf"

    from reportlab.lib.pagesizes import A4

    converter.convert_sheet_to_pdf("Sheet1", output_pdf, page_size=A4, orientation="landscape")

    assert output_pdf.exists()
    converter.close()


def test_convert_empty_sheet(sample_excel, tmp_path):
    """Test converting empty sheet raises error."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "empty.pdf"

    with pytest.raises(ValueError, match="empty"):
        converter.convert_sheet_to_pdf("EmptySheet", output_pdf)

    converter.close()


def test_convert_all_sheets(sample_excel, tmp_path):
    """Test converting all sheets to separate PDFs."""
    converter = ExcelToPDFConverter(sample_excel)
    output_dir = tmp_path / "output"

    created = converter.convert_all_sheets(output_dir)

    # Should create PDFs for non-empty sheets
    assert len(created) >= 2
    assert all(pdf.exists() for pdf in created)

    converter.close()


def test_convert_multiple_sheets(sample_excel, tmp_path):
    """Test converting multiple sheets to one PDF."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "combined.pdf"

    converter.convert_multiple_sheets(["Sheet1", "Sheet2"], output_pdf)

    assert output_pdf.exists()
    assert output_pdf.stat().st_size > 0

    converter.close()


def test_convert_with_letter_size(sample_excel, tmp_path):
    """Test conversion with Letter page size."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "letter.pdf"

    from reportlab.lib.pagesizes import letter

    converter.convert_sheet_to_pdf("Sheet1", output_pdf, page_size=letter)

    assert output_pdf.exists()
    converter.close()


def test_output_directory_creation(sample_excel, tmp_path):
    """Test that output directories are created automatically."""
    converter = ExcelToPDFConverter(sample_excel)
    output_pdf = tmp_path / "nested" / "dir" / "output.pdf"

    converter.convert_sheet_to_pdf("Sheet1", output_pdf)

    assert output_pdf.exists()
    assert output_pdf.parent.exists()

    converter.close()


def test_special_characters_in_sheet_name(tmp_path):
    """Test handling sheets with special characters."""
    excel_file = tmp_path / "special.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet (1) & 2"
    ws.append(["A", "B", "C"])
    ws.append(["1", "2", "3"])
    wb.save(excel_file)

    converter = ExcelToPDFConverter(excel_file)
    output_dir = tmp_path / "output"

    created = converter.convert_all_sheets(output_dir)

    assert len(created) == 1
    assert created[0].exists()

    converter.close()


def test_large_dataset(tmp_path):
    """Test conversion with larger dataset."""
    excel_file = tmp_path / "large.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LargeData"

    # Header
    ws.append(["Col1", "Col2", "Col3", "Col4", "Col5"])

    # 100 rows of data
    for i in range(100):
        ws.append([f"Data{i}", i, i * 2, i * 3, i * 4])

    wb.save(excel_file)

    converter = ExcelToPDFConverter(excel_file)
    output_pdf = tmp_path / "large.pdf"

    converter.convert_sheet_to_pdf("LargeData", output_pdf)

    assert output_pdf.exists()
    assert output_pdf.stat().st_size > 0

    converter.close()


def test_multiple_conversions(sample_excel, tmp_path):
    """Test multiple conversions with same converter."""
    converter = ExcelToPDFConverter(sample_excel)

    pdf1 = tmp_path / "output1.pdf"
    pdf2 = tmp_path / "output2.pdf"

    converter.convert_sheet_to_pdf("Sheet1", pdf1)
    converter.convert_sheet_to_pdf("Sheet2", pdf2)

    assert pdf1.exists()
    assert pdf2.exists()

    converter.close()


def test_close_converter(sample_excel):
    """Test closing converter."""
    converter = ExcelToPDFConverter(sample_excel)
    assert converter.workbook is not None
    converter.close()

    # Workbook is closed successfully
    # Note: openpyxl doesn't raise an error on closed workbook access
    # Just verify the close method runs without errors
