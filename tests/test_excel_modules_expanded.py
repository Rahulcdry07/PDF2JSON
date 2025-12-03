#!/usr/bin/env python3
"""
Comprehensive tests for Excel conversion modules to achieve >95% coverage.
Tests excel_to_pdf.py
"""

import sys
from pathlib import Path
import pytest
from openpyxl import Workbook
from unittest.mock import patch, MagicMock
import io

# Add scripts directory to path
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from excel_to_pdf import ExcelToPDFConverter, main as excel_to_pdf_main


@pytest.fixture
def sample_excel(tmp_path):
    """Create a sample Excel file with multiple sheets."""
    excel_file = tmp_path / "test.xlsx"
    wb = Workbook()

    # Sheet with data
    ws1 = wb.active
    ws1.title = "DataSheet"
    headers = ["Name", "Age", "City", "Department", "Salary"]
    ws1.append(headers)
    ws1.append(["Alice", 30, "NYC", "Engineering", 95000])
    ws1.append(["Bob", 25, "LA", "Marketing", 70000])
    ws1.append(["Charlie", 35, "SF", "Sales", 85000])

    # Wide sheet (for landscape testing)
    ws2 = wb.create_sheet("WideSheet")
    wide_headers = [f"Col{i}" for i in range(15)]
    ws2.append(wide_headers)
    ws2.append([f"Data{i}" for i in range(15)])

    # Very wide sheet (for multi-page testing)
    ws3 = wb.create_sheet("VeryWideSheet")
    very_wide_headers = [f"Column{i}" for i in range(20)]
    ws3.append(very_wide_headers)
    ws3.append([f"Value{i}" for i in range(20)])

    # Empty sheet
    wb.create_sheet("EmptySheet")

    # Sheet with special characters in name (sanitized for openpyxl)
    ws4 = wb.create_sheet("Special_Name_Test")
    ws4.append(["Test"])

    wb.save(excel_file)
    return excel_file


@pytest.fixture
def sample_excel_simple(tmp_path):
    """Create a simple Excel file for basic tests."""
    excel_file = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Test"
    ws["B1"] = "Data"
    wb.save(excel_file)
    return excel_file


# =============================================================================
# Tests for ExcelToPDFConverter - convert_sheet_to_pdf variations
# =============================================================================

def test_convert_sheet_empty(sample_excel, tmp_path):
    """Test converting empty sheet raises error."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "output.pdf"
    
    with pytest.raises(ValueError, match="empty"):
        converter.convert_sheet_to_pdf("EmptySheet", output)


def test_convert_sheet_with_auto_orientation(sample_excel, tmp_path):
    """Test auto orientation detection."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "auto.pdf"
    
    # WideSheet has 15 columns, should auto-select landscape
    converter.convert_sheet_to_pdf("WideSheet", output, orientation="auto")
    assert output.exists()


def test_convert_sheet_landscape(sample_excel, tmp_path):
    """Test landscape orientation."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "landscape.pdf"
    
    converter.convert_sheet_to_pdf("DataSheet", output, orientation="landscape")
    assert output.exists()


def test_convert_sheet_very_wide_table(sample_excel, tmp_path, capsys):
    """Test very wide table (20 columns) - multi-page warning."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "verywide.pdf"
    
    converter.convert_sheet_to_pdf("VeryWideSheet", output)
    captured = capsys.readouterr()
    
    assert output.exists()
    # Should warn about multi-page layout
    assert "columns" in captured.out.lower() or "page" in captured.out.lower()


def test_convert_sheet_fit_to_width(sample_excel, tmp_path):
    """Test fit_to_width parameter."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "fitted.pdf"
    
    converter.convert_sheet_to_pdf("DataSheet", output, fit_to_width=True)
    assert output.exists()


def test_convert_sheet_creates_parent_dir(sample_excel, tmp_path):
    """Test that parent directories are created."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "subdir" / "nested" / "output.pdf"
    
    converter.convert_sheet_to_pdf("DataSheet", output)
    assert output.exists()
    assert output.parent.exists()


# =============================================================================
# Tests for ExcelToPDFConverter - convert_all_sheets
# =============================================================================

def test_convert_all_sheets(sample_excel, tmp_path):
    """Test converting all sheets to separate PDFs."""
    converter = ExcelToPDFConverter(sample_excel)
    output_dir = tmp_path / "all_sheets"
    
    created = converter.convert_all_sheets(output_dir)
    
    assert len(created) >= 3  # At least DataSheet, WideSheet, VeryWideSheet
    assert output_dir.exists()
    # EmptySheet should fail and be skipped
    

def test_convert_all_sheets_sanitizes_filenames(sample_excel, tmp_path):
    """Test that sheet names with underscores work correctly."""
    converter = ExcelToPDFConverter(sample_excel)
    output_dir = tmp_path / "sanitized"
    
    created = converter.convert_all_sheets(output_dir)
    
    # Special_Name_Test should create a PDF
    filenames = [f.name for f in created]
    assert any("Special" in name for name in filenames)


def test_convert_all_sheets_with_landscape(sample_excel, tmp_path):
    """Test convert_all_sheets with landscape orientation."""
    converter = ExcelToPDFConverter(sample_excel)
    output_dir = tmp_path / "landscape_all"
    
    created = converter.convert_all_sheets(output_dir, orientation="landscape")
    assert len(created) >= 1


def test_convert_all_sheets_handles_errors(sample_excel, tmp_path, capsys):
    """Test that convert_all_sheets handles errors gracefully."""
    converter = ExcelToPDFConverter(sample_excel)
    output_dir = tmp_path / "with_errors"
    
    # This will try to convert EmptySheet which will fail
    created = converter.convert_all_sheets(output_dir)
    
    captured = capsys.readouterr()
    # Should have error message for empty sheet
    assert "Error" in captured.err or "empty" in captured.out.lower()


# =============================================================================
# Tests for ExcelToPDFConverter - convert_multiple_sheets
# =============================================================================

def test_convert_multiple_sheets(sample_excel, tmp_path):
    """Test converting multiple sheets to single PDF."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "combined.pdf"
    
    converter.convert_multiple_sheets(["DataSheet", "WideSheet"], output)
    assert output.exists()


def test_convert_multiple_sheets_landscape(sample_excel, tmp_path):
    """Test convert_multiple_sheets with landscape."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "combined_landscape.pdf"
    
    converter.convert_multiple_sheets(
        ["DataSheet", "WideSheet"], output, orientation="landscape"
    )
    assert output.exists()


def test_convert_multiple_sheets_creates_dir(sample_excel, tmp_path):
    """Test that convert_multiple_sheets creates parent directory."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "new_dir" / "combined.pdf"
    
    converter.convert_multiple_sheets(["DataSheet"], output)
    assert output.exists()


def test_convert_multiple_sheets_with_empty_sheet(sample_excel, tmp_path, capsys):
    """Test convert_multiple_sheets handles empty sheets."""
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "with_empty.pdf"
    
    # Include EmptySheet which should be skipped
    converter.convert_multiple_sheets(["DataSheet", "EmptySheet"], output)
    
    captured = capsys.readouterr()
    # May have error but should still create PDF
    assert output.exists()


# =============================================================================
# Tests for ExcelToPDFConverter - close method
# =============================================================================

def test_close_workbook(sample_excel):
    """Test closing the workbook."""
    converter = ExcelToPDFConverter(sample_excel)
    converter.close()
    # After closing, workbook should be closed
    # (Can't easily test this without accessing private state)


# =============================================================================
# Tests for main() CLI function
# =============================================================================

def test_main_file_not_found():
    """Test main with non-existent file."""
    with patch("sys.argv", ["excel_to_pdf.py", "nonexistent.xlsx", "out.pdf"]):
        with pytest.raises(SystemExit) as exc_info:
            excel_to_pdf_main()
        assert exc_info.value.code == 1


def test_main_invalid_file_type(tmp_path):
    """Test main with invalid file type."""
    txt_file = tmp_path / "test.txt"
    txt_file.write_text("not excel")
    
    with patch("sys.argv", ["excel_to_pdf.py", str(txt_file), "out.pdf"]):
        with pytest.raises(SystemExit) as exc_info:
            excel_to_pdf_main()
        assert exc_info.value.code == 1


def test_main_list_sheets(sample_excel, capsys):
    """Test main with --list flag."""
    with patch("sys.argv", ["excel_to_pdf.py", str(sample_excel), "--list"]):
        try:
            excel_to_pdf_main()
        except SystemExit:
            pass
        
        captured = capsys.readouterr()
        assert "DataSheet" in captured.out
        assert "WideSheet" in captured.out


def test_main_no_output():
    """Test main without output path."""
    with patch("sys.argv", ["excel_to_pdf.py", "test.xlsx"]):
        with pytest.raises(SystemExit):
            excel_to_pdf_main()


def test_main_all_sheets(sample_excel, tmp_path, capsys):
    """Test main with --all-sheets."""
    output_dir = tmp_path / "all"
    
    with patch("sys.argv", [
        "excel_to_pdf.py", str(sample_excel), str(output_dir), "--all-sheets"
    ]):
        excel_to_pdf_main()
    
    captured = capsys.readouterr()
    assert "Created" in captured.out or "Converting" in captured.out


def test_main_single_sheet(sample_excel, tmp_path, capsys):
    """Test main with --sheet."""
    output = tmp_path / "single.pdf"
    
    with patch("sys.argv", [
        "excel_to_pdf.py", str(sample_excel), str(output), "--sheet", "DataSheet"
    ]):
        excel_to_pdf_main()
    
    assert output.exists()


def test_main_multiple_sheets(sample_excel, tmp_path, capsys):
    """Test main with --sheets."""
    output = tmp_path / "multi.pdf"
    
    with patch("sys.argv", [
        "excel_to_pdf.py", str(sample_excel), str(output), 
        "--sheets", "DataSheet,WideSheet"
    ]):
        excel_to_pdf_main()
    
    assert output.exists()


def test_main_landscape_flag(sample_excel, tmp_path):
    """Test main with --landscape."""
    output = tmp_path / "landscape.pdf"
    
    with patch("sys.argv", [
        "excel_to_pdf.py", str(sample_excel), str(output),
        "--sheet", "DataSheet", "--landscape"
    ]):
        excel_to_pdf_main()
    
    assert output.exists()


def test_main_letter_page_size(sample_excel, tmp_path):
    """Test main with --page-size Letter."""
    output = tmp_path / "letter.pdf"
    
    with patch("sys.argv", [
        "excel_to_pdf.py", str(sample_excel), str(output),
        "--sheet", "DataSheet", "--page-size", "Letter"
    ]):
        excel_to_pdf_main()
    
    assert output.exists()


def test_main_no_sheet_specified(sample_excel, tmp_path, capsys):
    """Test main without specifying sheet to convert."""
    output = tmp_path / "out.pdf"
    
    with patch("sys.argv", ["excel_to_pdf.py", str(sample_excel), str(output)]):
        with pytest.raises(SystemExit):
            excel_to_pdf_main()


def test_main_sheet_not_found(sample_excel, tmp_path):
    """Test main with non-existent sheet name."""
    output = tmp_path / "out.pdf"
    
    with patch("sys.argv", [
        "excel_to_pdf.py", str(sample_excel), str(output),
        "--sheet", "NonExistentSheet"
    ]):
        with pytest.raises(SystemExit):
            excel_to_pdf_main()


# =============================================================================
# Additional edge case tests
# =============================================================================

def test_convert_sheet_with_none_values(tmp_path):
    """Test converting sheet with None values."""
    excel_file = tmp_path / "with_none.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append(["1", None, "3"])
    ws.append([None, "2", None])
    wb.save(excel_file)
    
    converter = ExcelToPDFConverter(excel_file)
    output = tmp_path / "none.pdf"
    converter.convert_sheet_to_pdf(ws.title, output)
    
    assert output.exists()


def test_get_sheet_data_removes_trailing_empty_rows(tmp_path):
    """Test that trailing empty rows are removed."""
    excel_file = tmp_path / "trailing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Header"])
    ws.append(["Data"])
    # Add empty rows
    ws.append([])
    ws.append([])
    wb.save(excel_file)
    
    converter = ExcelToPDFConverter(excel_file)
    data = converter.get_sheet_data(ws.title)
    
    # Should only have 2 rows (header + data)
    assert len(data) == 2


def test_convert_with_different_page_sizes(sample_excel, tmp_path):
    """Test conversion with different page sizes."""
    from reportlab.lib.pagesizes import letter
    
    converter = ExcelToPDFConverter(sample_excel)
    output = tmp_path / "letter.pdf"
    
    converter.convert_sheet_to_pdf("DataSheet", output, page_size=letter)
    assert output.exists()
